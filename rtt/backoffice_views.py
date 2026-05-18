# Backoffice: dashboard espelho de ponto, aprovação, indicadores e exportação
from datetime import datetime, timedelta, time
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods, require_GET, require_POST
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Count, Q, Sum
from django.db.models.functions import TruncDate

from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout

from .models import Marcacao, Profile, Jornada, Departamento, RegistroKM, Viatura

from .filters import EspelhoPontoFilter

User = get_user_model()

# Rótulos de marcação alinhados à área do utilizador (1ª a 4ª marcação)
TIPO_MARCACAO_LABEL = {
    'entrada': '1ª marcação',
    'inicio_almoco': '2ª marcação',
    'fim_almoco': '3ª marcação',
    'fim_jornada': '4ª marcação',
}

TIPO_FRIENDLY_LABEL = {
    'entrada': 'Entrada',
    'inicio_almoco': 'Início Almoço',
    'fim_almoco': 'Volta Almoço',
    'fim_jornada': 'Saída',
}

BACKOFFICE_LOGIN_URL = '/'


def backoffice_required(view_func):
    """Decorator: exige login e is_staff. Redireciona para tela de login unificada."""
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from urllib.parse import quote
            next_url = quote(request.get_full_path(), safe='')
            return redirect(BACKOFFICE_LOGIN_URL + '?next=' + next_url)
        if not request.user.is_staff:
            return redirect(BACKOFFICE_LOGIN_URL + '?erro=acesso_restrito')
        return view_func(request, *args, **kwargs)
    return _wrapped


def backoffice_login_view(request):
    """Redireciona para a tela de login unificada na raiz."""
    return redirect('/')


def backoffice_logout_view(request):
    """Termina apenas a sessão do backoffice e redireciona para o login do backoffice."""
    auth_logout(request)
    return redirect(BACKOFFICE_LOGIN_URL)


# ---------- Colaboradores ----------
@backoffice_required
def colaborador_list_view(request):
    """Lista colaboradores com link para novo e editar."""
    from django.db.models import Q
    perfis = Profile.objects.select_related('user', 'departamento', 'jornada').order_by('nome')
    q = request.GET.get('q')
    if q:
        perfis = perfis.filter(
            Q(nome__icontains=q) | Q(user__email__icontains=q)
        )
    return render(request, 'backoffice/colaborador_list.html', {'perfis': perfis})


@backoffice_required
def colaborador_create_view(request):
    """Formulário para cadastrar novo colaborador."""
    from .forms import ColaboradorForm
    if request.method == 'POST':
        form = ColaboradorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('backoffice_colaborador_list')
    else:
        form = ColaboradorForm()
    return render(request, 'backoffice/colaborador_form.html', {'form': form, 'titulo': 'Novo colaborador'})


@backoffice_required
def colaborador_edit_view(request, pk):
    """Formulário para editar colaborador."""
    from .forms import ColaboradorForm
    profile = get_object_or_404(Profile, pk=pk)
    if request.method == 'POST':
        form = ColaboradorForm(request.POST, instance=profile, edit_user=profile.user)
        if form.is_valid():
            form.save()
            return redirect('backoffice_colaborador_list')
    else:
        form = ColaboradorForm(instance=profile, edit_user=profile.user)
    return render(request, 'backoffice/colaborador_form.html', {'form': form, 'titulo': 'Editar colaborador', 'profile': profile})

# ---------- Departamentos ----------
@backoffice_required
def departamento_list_view(request):
    """Lista departamentos (origem do dropdown Departamento no cadastro de colaborador)."""
    lista = Departamento.objects.all().order_by('nome')
    return render(request, 'backoffice/departamento_list.html', {'lista': lista})


@backoffice_required
def departamento_create_view(request):
    """Criar departamento."""
    from .forms import DepartamentoForm
    if request.method == 'POST':
        form = DepartamentoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('backoffice_departamento_list')
    else:
        form = DepartamentoForm()
    return render(request, 'backoffice/departamento_form.html', {'form': form, 'titulo': 'Novo departamento'})


@backoffice_required
def departamento_edit_view(request, pk):
    """Editar departamento."""
    from .forms import DepartamentoForm
    obj = get_object_or_404(Departamento, pk=pk)
    if request.method == 'POST':
        form = DepartamentoForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('backoffice_departamento_list')
    else:
        form = DepartamentoForm(instance=obj)
    return render(request, 'backoffice/departamento_form.html', {'form': form, 'titulo': 'Editar departamento', 'obj': obj})


# ---------- Jornadas ----------
@backoffice_required
def jornada_list_view(request):
    """Lista jornadas (origem do dropdown Jornada no cadastro de colaborador)."""
    lista = Jornada.objects.all().order_by('nome')
    return render(request, 'backoffice/jornada_list.html', {'lista': lista})


@backoffice_required
def jornada_create_view(request):
    """Criar jornada."""
    from .forms import JornadaForm
    if request.method == 'POST':
        form = JornadaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('backoffice_jornada_list')
    else:
        form = JornadaForm()
    return render(request, 'backoffice/jornada_form.html', {'form': form, 'titulo': 'Nova jornada'})


@backoffice_required
def jornada_edit_view(request, pk):
    """Editar jornada."""
    from .forms import JornadaForm
    obj = get_object_or_404(Jornada, pk=pk)
    if request.method == 'POST':
        form = JornadaForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('backoffice_jornada_list')
    else:
        form = JornadaForm(instance=obj)
    return render(request, 'backoffice/jornada_form.html', {'form': form, 'titulo': 'Editar jornada', 'obj': obj})


@backoffice_required
def colaborador_detail_view(request, pk):
    """Detalhe do colaborador: ficha e marcações com localização GPS."""
    profile = get_object_or_404(Profile.objects.select_related('user', 'departamento', 'jornada'), pk=pk)
    # filtros simples por data
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    qs = Marcacao.objects.filter(utilizador=profile.user).order_by('-timestamp')
    if data_inicio:
        try:
            di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            qs = qs.filter(timestamp__date__gte=di)
        except ValueError:
            data_inicio = ''
    if data_fim:
        try:
            df = datetime.strptime(data_fim, '%Y-%m-%d').date()
            qs = qs.filter(timestamp__date__lte=df)
        except ValueError:
            data_fim = ''
    # Estrutura "espelho": 1 linha por dia, até 8 marcações (por ordem cronológica)
    from collections import OrderedDict
    por_dia = OrderedDict()
    # Para o "espelho", a ordem dentro do dia deve ser cronológica (asc)
    for m in qs.order_by('timestamp')[:500]:
        ts = timezone.localtime(m.timestamp)
        dia = ts.date()
        if dia not in por_dia:
            wd = ts.weekday()
            por_dia[dia] = {
                'data': dia,
                'dia_semana': ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom'][wd],
                'marcacoes': []
            }
        if len(por_dia[dia]['marcacoes']) >= 4:
            continue
        lat = str(m.latitude)
        lon = str(m.longitude)
        por_dia[dia]['marcacoes'].append({
            'id': m.pk,
            'hora': ts.strftime('%H:%M'),
            'tipo': TIPO_MARCACAO_LABEL.get(m.tipo, m.get_tipo_display()),
            'tipo_raw': m.tipo,
            'latitude': lat,
            'longitude': lon,
            'mapa_url': f'https://www.google.com/maps?q={lat},{lon}',
            'tem_mapa': (m.latitude and m.longitude and float(m.latitude) != 0.0 and float(m.longitude) != 0.0),
        })

    # Mais recente em cima
    dias = list(reversed(list(por_dia.values())))
    for d in dias:
        # preencher até 8 células vazias para manter layout
        while len(d['marcacoes']) < 4:
            d['marcacoes'].append({'hora': '—', 'tipo': '', 'latitude': '0', 'longitude': '0', 'mapa_url': '', 'tem_mapa': False})
    return render(request, 'backoffice/colaborador_detail.html', {
        'profile': profile,
        'dias': dias,
        'data_inicio': data_inicio or '',
        'data_fim': data_fim or '',
    })


def _get_jornada_planejada(profile):
    """Retorna (h_entrada, h_saida, h_ini_int, h_fim_int) em time. Default 08:00-18:00, 12:00-13:00."""
    if profile and profile.jornada_id:
        j = profile.jornada
        return (
            j.hora_entrada_planejada or time(8, 0),
            j.hora_saida_planejada or time(18, 0),
            j.hora_inicio_intervalo or time(12, 0),
            j.hora_fim_intervalo or time(13, 0),
        )
    return (time(8, 0), time(18, 0), time(12, 0), time(13, 0))


def _marcacoes_por_dia(utilizador_id, data_inicio, data_fim):
    """Retorna dict[date_str -> dict[tipo -> {time, aprovado}]] para o utilizador no intervalo."""
    qs = Marcacao.objects.filter(
        utilizador_id=utilizador_id,
        timestamp__date__gte=data_inicio,
        timestamp__date__lte=data_fim,
    ).order_by('timestamp')
    from collections import defaultdict
    por_dia = defaultdict(dict)
    for m in qs:
        dt = timezone.localtime(m.timestamp)
        dia = dt.date().isoformat()
        # Fica a última marcação do dia para cada tipo (em caso de duplicados)
        por_dia[dia][m.tipo] = {'time': dt.strftime('%H:%M'), 'aprovado': m.aprovado}
    return por_dia


def _marcacoes_ordenadas_por_dia(utilizador_id, data_inicio, data_fim):
    """Retorna dict[date_str -> list de até 8 horários (HH:MM)] por ordem cronológica."""
    qs = Marcacao.objects.filter(
        utilizador_id=utilizador_id,
        timestamp__date__gte=data_inicio,
        timestamp__date__lte=data_fim,
    ).order_by('timestamp')
    from collections import defaultdict
    por_dia = defaultdict(list)
    for m in qs:
        dt = timezone.localtime(m.timestamp)
        dia = dt.date().isoformat()
        if len(por_dia[dia]) < 4:
            por_dia[dia].append(dt.strftime('%H:%M'))
    return por_dia


def _status_celula(hora_registada_str, hora_planejada, tipo_compare):
    """
    tipo_compare: 'entrada' (atraso se reg > planejado), 'saida' (atraso se reg < planejado),
    'intervalo' (só falta/ok). Retorna 'ok' (verde), 'atraso'/'falta' (vermelho), 'pendente' (amarelo).
    """
    if not hora_planejada:
        return 'ok' if hora_registada_str else 'falta'
    if not hora_registada_str:
        return 'falta'
    try:
        h, m = map(int, hora_registada_str.split(':'))
        reg = time(h, m)
    except (ValueError, TypeError):
        return 'ok'
    planejado = hora_planejada if isinstance(hora_planejada, time) else hora_planejada
    if tipo_compare == 'entrada':
        return 'atraso' if reg > planejado else 'ok'
    if tipo_compare == 'saida':
        return 'atraso' if reg < planejado else 'ok'
    return 'ok'


def _calcular_total(entrada, saida, inicio_int, fim_int):
    """Calcula minutos trabalhados (entrada-saída menos intervalo). Retorna string HH:MM ou None."""
    def to_minutes(t):
        if not t:
            return None
        if isinstance(t, time):
            return t.hour * 60 + t.minute
        try:
            h, m = map(int, str(t).split(':')[:2])
            return h * 60 + m
        except (ValueError, TypeError):
            return None
    m_ent = to_minutes(entrada)
    m_sai = to_minutes(saida)
    if m_ent is None or m_sai is None:
        return None
    dur = m_sai - m_ent
    if inicio_int and fim_int:
        m_ini = to_minutes(inicio_int)
        m_fim = to_minutes(fim_int)
        if m_ini is not None and m_fim is not None:
            dur -= (m_fim - m_ini)
    if dur < 0:
        return None
    return f'{dur // 60:02d}:{dur % 60:02d}'


def _calcular_total_por_pares(lista_horas):
    """
    Calcula o total trabalhado a partir de uma sequência ordenada de horas (HH:MM), Entrada até Saída.
    Regra: soma todos os intervalos consecutivos (2ª-1ª)+(3ª-2ª)+(4ª-3ª)
    = tempo da primeira à última marcação. Assim usa todas as marcações de 1 a 4.
    Retorna string HH:MM ou None.
    """
    def to_minutes(t):
        if not t:
            return None
        if isinstance(t, time):
            return t.hour * 60 + t.minute
        try:
            h, m = map(int, str(t).split(':')[:2])
            return h * 60 + m
        except (ValueError, TypeError):
            return None

    if not lista_horas:
        return None
    mins = []
    for h in lista_horas:
        m = to_minutes(h)
        if m is not None:
            mins.append(m)
    if len(mins) < 2:
        return None

    total = 0
    for i in range(len(mins) - 1):
        dur = mins[i + 1] - mins[i]
        if dur < 0:
            return None
        total += dur
    return f'{total // 60:02d}:{total % 60:02d}'


def construir_espelho(data_inicio, data_fim, utilizadores_queryset):
    """
    Constrói lista de linhas do espelho de ponto: uma linha por (utilizador, dia).
    Cada linha: utilizador, nome, data, jornada_nome, planejado_str, entrada, entrada_status, ...
    """
    from datetime import date
    linhas = []
    delta = (data_fim - data_inicio).days + 1
    for user in utilizadores_queryset:
        try:
            profile = user.profile
            nome = profile.nome or user.email or str(user.pk)
            jornada = profile.jornada
            jornada_nome = jornada.nome if jornada else 'Padrão 08h-18h'
            h_ent, h_sai, h_ini, h_fim = _get_jornada_planejada(profile)
            planejado_str = h_ent.strftime('%H:%M')
        except Profile.DoesNotExist:
            nome = user.email or str(user.pk)
            jornada_nome = 'Padrão 08h-18h'
            planejado_str = '08:00'
            h_ent, h_sai, h_ini, h_fim = time(8, 0), time(18, 0), time(12, 0), time(13, 0)

        for i in range(delta):
            dia = data_inicio + timedelta(days=i)
            qs_dia = Marcacao.objects.filter(utilizador=user, timestamp__date=dia).order_by('timestamp')
            if not qs_dia.exists():
                continue

            detalhes_marcs = []
            horas_sozinhos = []
            for m in qs_dia:
                h_str = timezone.localtime(m.timestamp).strftime('%H:%M')
                lat = str(m.latitude)
                lon = str(m.longitude)
                detalhes_marcs.append({
                    'hora': h_str,
                    'tipo': m.tipo,
                    'tipo_display': TIPO_FRIENDLY_LABEL.get(m.tipo, m.get_tipo_display()),
                    'latitude': lat,
                    'longitude': lon,
                    'mapa_url': f'https://www.google.com/maps?q={lat},{lon}',
                    'tem_mapa': (m.latitude and m.longitude and float(m.latitude) != 0.0 and float(m.longitude) != 0.0),
                })
                horas_sozinhos.append(h_str)

            total = _calcular_total_por_pares(horas_sozinhos)
            wd = dia.weekday()
            is_weekend = wd >= 5

            linhas.append({
                'utilizador': user,
                'nome': nome,
                'data': dia,
                'data_str': dia.strftime('%d/%m/%Y'),
                'dia_semana': ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom'][wd],
                'jornada_nome': jornada_nome,
                'planejado': planejado_str,
                'marcacoes': detalhes_marcs,
                'total': total or '—',
                'is_weekend': is_weekend,
            })
    
    # Ordenação (mais recente primeiro)
    linhas.sort(key=lambda x: (x['data'], (x['nome'] or '').lower()), reverse=True)
    return linhas


@backoffice_required
def backoffice_km_list_view(request):
    """Lista de registos de KM com totais."""
    registos = RegistroKM.objects.select_related('utilizador', 'utilizador__profile').order_by('-timestamp')
    
    colaborador_id = request.GET.get('colaborador')
    if colaborador_id:
        registos = registos.filter(utilizador_id=colaborador_id)
        
    data_inicio = request.GET.get('data_inicio')
    if data_inicio:
        registos = registos.filter(data__gte=data_inicio)
        
    data_fim = request.GET.get('data_fim')
    if data_fim:
        registos = registos.filter(data__lte=data_fim)

    viatura_id = request.GET.get('viatura')
    if viatura_id:
        registos = registos.filter(viatura_id=viatura_id)

    # Cálculo do total do período filtrado (Soma do deslocamento diário por viatura/utilizador)
    from django.db.models import Max, Min
    resumo = registos.order_by().values('data', 'viatura', 'utilizador').annotate(
        max_km=Max('km'),
        min_prev=Min('km_anterior')
    )
    total_periodo = 0
    for r in resumo:
        if r['max_km'] is not None and r['min_prev'] is not None:
            diff = r['max_km'] - r['min_prev']
            if diff > 0:
                total_periodo += diff

    # Totais específicos se houver colaborador selecionado
    stats_colaborador = None
    if colaborador_id:
        from django.utils import timezone
        hoje = timezone.localdate()
        
        def get_total_range(start_date):
            qs = RegistroKM.objects.filter(utilizador_id=colaborador_id, data__gte=start_date)
            res = qs.order_by().values('data', 'viatura').annotate(
                max_km=Max('km'),
                min_prev=Min('km_anterior')
            )
            s = 0
            for r in res:
                if r['max_km'] is not None and r['min_prev'] is not None:
                    d = r['max_km'] - r['min_prev']
                    if d > 0: s += d
            return float(s)

        stats_colaborador = {
            'hoje': get_total_range(hoje),
            'semana': get_total_range(hoje - timezone.timedelta(days=hoje.weekday())),
            'mes': get_total_range(hoje.replace(day=1))
        }

    colaboradores = User.objects.filter(profile__isnull=False).distinct().order_by('profile__nome')
    viaturas = Viatura.objects.filter(ativo=True).order_by('matricula')

    # Ajustes se houver colaborador selecionado
    if colaborador_id:
        # Filtrar viaturas: apenas as que têm registos deste colaborador
        viaturas = viaturas.filter(registos__utilizador_id=colaborador_id).distinct()
        
        # Se não houver data_inicio, buscar a data do 1º registo do colaborador
        if not data_inicio:
            primeiro_registo = RegistroKM.objects.filter(utilizador_id=colaborador_id).order_by('data').first()
            if primeiro_registo:
                data_inicio = primeiro_registo.data.strftime('%Y-%m-%d')
                # Re-filtrar os registos com a nova data_inicio
                registos = registos.filter(data__gte=primeiro_registo.data)

    return render(request, 'backoffice/km_list.html', {
        'registos': registos,
        'colaboradores': colaboradores,
        'viaturas': viaturas,
        'colaborador_selecionado': colaborador_id,
        'viatura_selecionada': viatura_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'total_periodo': float(total_periodo),
        'stats_colaborador': stats_colaborador
    })


@backoffice_required
def backoffice_km_create_view(request):
    """Permite ao administrador registrar um KM manualmente."""
    from .forms import RegistroKMForm
    if request.method == 'POST':
        form = RegistroKMForm(request.POST)
        if form.is_valid():
            registo = form.save(commit=False)
            viatura = registo.viatura
            # Validação: KM deve ser superior ao atual da viatura
            if registo.km < viatura.km_atual:
                form.add_error('km', f'O KM não pode ser menor que o atual da viatura ({viatura.km_atual}).')
            else:
                registo.km_anterior = viatura.km_atual
                registo.save()
                viatura.km_atual = registo.km
                viatura.save()
                return redirect('backoffice_km_list')
    else:
        form = RegistroKMForm()
    return render(request, 'backoffice/km_form.html', {'form': form, 'titulo': 'Lançar Registro de KM'})


@backoffice_required
@require_POST
def backoffice_km_delete_view(request, pk):
    """Eliminar registo de KM com confirmação por senha (senha: 1143)."""
    registo = get_object_or_404(RegistroKM, pk=pk)
    senha_inserida = request.POST.get('senha', '').strip()
    SENHA_ELIMINACAO = '1143'

    # Construir URL de retorno com os filtros anteriores se existirem
    referer = request.POST.get('referer', '')
    redirect_url = referer if referer else '/backoffice/km-registros/'

    if senha_inserida != SENHA_ELIMINACAO:
        # Senha errada — redirecionar com erro
        from urllib.parse import urlencode, urlparse, parse_qs, urlunparse
        parsed = urlparse(redirect_url)
        params = parse_qs(parsed.query)
        params['erro_senha'] = ['1']
        new_query = urlencode({k: v[0] for k, v in params.items()})
        return redirect(urlunparse(parsed._replace(query=new_query)))

    # Senha correta — eliminar o registo
    registo.delete()

    # Redirecionar com mensagem de sucesso
    from urllib.parse import urlencode, urlparse, parse_qs, urlunparse
    parsed = urlparse(redirect_url)
    params = parse_qs(parsed.query)
    params.pop('erro_senha', None)
    params['eliminado'] = ['1']
    new_query = urlencode({k: v[0] for k, v in params.items()})
    return redirect(urlunparse(parsed._replace(query=new_query)))


@backoffice_required
@require_GET
def dashboard_espelho_view(request):
    """Dashboard espelho de ponto: tabela Data, Jornada, Planejado, Entrada, Início Int., Volta Int., Saída, Total."""
    colaboradores = User.objects.filter(profile__isnull=False).distinct().order_by('profile__nome')
    if not colaboradores.exists():
        colaboradores = User.objects.all().order_by('email')[:50]

    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    utilizador_id = request.GET.get('colaborador')
    today = timezone.localdate()
    if not data_inicio:
        # Semana atual (segunda a domingo)
        wd = today.weekday()
        seg = today - timedelta(days=wd)
        data_inicio = seg
    else:
        try:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        except ValueError:
            data_inicio = today - timedelta(days=today.weekday())
    if not data_fim:
        data_fim = data_inicio + timedelta(days=6)
    else:
        try:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        except ValueError:
            data_fim = data_inicio + timedelta(days=6)

    if utilizador_id:
        colaboradores = colaboradores.filter(pk=utilizador_id)
    if not colaboradores.exists():
        colaboradores = User.objects.none()

    linhas = construir_espelho(data_inicio, data_fim, colaboradores)
    filter_form = EspelhoPontoFilter(request.GET, queryset=Marcacao.objects.none())

    return render(request, 'backoffice/dashboard_espelho.html', {
        'linhas': linhas,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'colaboradores': User.objects.filter(profile__isnull=False).distinct().order_by('profile__nome'),
        'filter_form': filter_form,
        'colaborador_selecionado': utilizador_id,
    })


@backoffice_required
@require_POST
@csrf_exempt
def aprovar_ponto_view(request):
    """Aprova uma marcação (ou várias do mesmo dia). POST: marcacao_id ou utilizador_id + data + tipo."""
    from django.utils import timezone as tz
    marcacao_id = request.POST.get('marcacao_id')
    if marcacao_id:
        marcacao = get_object_or_404(Marcacao, id=marcacao_id)
        marcacao.aprovado = True
        marcacao.aprovado_por = request.user
        marcacao.aprovado_em = tz.now()
        marcacao.save()
        return JsonResponse({'sucesso': True, 'mensagem': 'Ponto aprovado.'})
    # Aprovar todos do dia para um tipo
    utilizador_id = request.POST.get('utilizador_id')
    data_str = request.POST.get('data')
    tipo = request.POST.get('tipo')
    if not utilizador_id or not data_str:
        return JsonResponse({'sucesso': False, 'erro': 'utilizador_id e data obrigatórios.'}, 400)
    try:
        data = datetime.strptime(data_str, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'sucesso': False, 'erro': 'Data inválida.'}, 400)
    qs = Marcacao.objects.filter(utilizador_id=utilizador_id, timestamp__date=data)
    if tipo:
        qs = qs.filter(tipo=tipo)
    count = qs.count()
    qs.update(aprovado=True, aprovado_por=request.user, aprovado_em=tz.now())
    return JsonResponse({'sucesso': True, 'mensagem': 'Pontos aprovados.', 'count': count})


@backoffice_required
@require_POST
@csrf_exempt
def adicionar_ponto_view(request):
    """Adiciona uma marcação manualmente (admin). POST: utilizador_id, data, hora, tipo_marcacao."""
    from decimal import Decimal
    utilizador_id = request.POST.get('utilizador_id')
    data_str = request.POST.get('data')
    hora_str = request.POST.get('hora')
    tipo = (request.POST.get('tipo_marcacao') or '').strip().lower()
    if tipo not in ('entrada', 'inicio_almoco', 'fim_almoco', 'fim_jornada'):
        return JsonResponse({'sucesso': False, 'erro': 'tipo_marcacao inválido.'}, 400)
    if not utilizador_id or not data_str or not hora_str:
        return JsonResponse({'sucesso': False, 'erro': 'utilizador_id, data e hora obrigatórios.'}, 400)
    try:
        dt = datetime.strptime(data_str + ' ' + hora_str, '%Y-%m-%d %H:%M')
    except ValueError:
        return JsonResponse({'sucesso': False, 'erro': 'Data/hora inválidos. Use YYYY-MM-DD e HH:MM.'}, 400)
    user = get_object_or_404(User, pk=utilizador_id)
    dt_aware = timezone.make_aware(dt)
    dia = dt_aware.date()
    if Marcacao.objects.filter(utilizador=user, timestamp__date=dia).count() >= 8:
        return JsonResponse({'sucesso': False, 'erro': 'Limite de 8 marcações por dia atingido para este colaborador.'}, 400)
    Marcacao.objects.create(
        utilizador=user,
        tipo=tipo,
        latitude=Decimal('0'),
        longitude=Decimal('0'),
        timestamp=dt_aware,
        aprovado=True,
        aprovado_por=request.user,
        aprovado_em=timezone.now(),
    )
    return JsonResponse({'sucesso': True, 'mensagem': 'Ponto adicionado.'})


@backoffice_required
def ponto_edit_view(request, pk):
    """Permite ao administrador editar uma marcação de ponto específica."""
    from .forms import MarcacaoForm
    marcacao = get_object_or_404(Marcacao, pk=pk)
    colaborador_pk = marcacao.utilizador.profile.pk
    if request.method == 'POST':
        form = MarcacaoForm(request.POST, instance=marcacao)
        if form.is_valid():
            form.save()
            return redirect('backoffice_colaborador_detail', pk=colaborador_pk)
    else:
        form = MarcacaoForm(instance=marcacao)
    return render(request, 'backoffice/ponto_form.html', {
        'form': form, 
        'titulo': 'Editar Marcação', 
        'marcacao': marcacao,
        'colaborador_pk': colaborador_pk
    })


@backoffice_required
@require_POST
def ponto_delete_view(request, pk):
    """Permite ao administrador excluir uma marcação de ponto."""
    marcacao = get_object_or_404(Marcacao, pk=pk)
    colaborador_pk = marcacao.utilizador.profile.pk
    marcacao.delete()
    return redirect('backoffice_colaborador_detail', pk=colaborador_pk)


@backoffice_required
@require_GET
def indicadores_view(request):
    """Indicadores BI: absenteísmo por departamento, alertas de inconsistência."""
    hoje = timezone.localdate()
    inicio_mes = hoje.replace(day=1)
    dept_absenteismo = {}
    for p in Profile.objects.select_related('departamento', 'user').all():
        dnome = p.departamento.nome if p.departamento else 'Sem departamento'
        if dnome not in dept_absenteismo:
            dept_absenteismo[dnome] = {'colaboradores': 0, 'dias_com_entrada': 0}
        dept_absenteismo[dnome]['colaboradores'] += 1
        entradas = Marcacao.objects.filter(utilizador=p.user, tipo='entrada', timestamp__date__gte=inicio_mes, timestamp__date__lte=hoje).values_list('timestamp', flat=True)
        dias_com = len(set(timezone.localtime(ts).date() for ts in entradas))
        dept_absenteismo[dnome]['dias_com_entrada'] += dias_com
    alertas = []
    for p in Profile.objects.select_related('user').all():
        ultimas = Marcacao.objects.filter(utilizador=p.user).order_by('-timestamp')[:20]
        for m in ultimas:
            dt = timezone.localtime(m.timestamp).date()
            if m.tipo == 'entrada':
                tem_saida = Marcacao.objects.filter(utilizador=p.user, tipo='fim_jornada', timestamp__date=dt).exists()
                if not tem_saida and dt >= hoje - timedelta(days=7):
                    alertas.append({'nome': p.nome or p.user.email, 'data': dt, 'tipo': 'Sem marcação de saída'})
                break
    return render(request, 'backoffice/indicadores.html', {
        'dept_absenteismo': dept_absenteismo,
        'alertas': alertas[:30],
    })


@backoffice_required
@require_GET
def export_espelho_excel_view(request):
    """Exporta espelho de ponto (intervalo) para Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return HttpResponse('openpyxl não instalado. pip install openpyxl', status=500)
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    today = timezone.localdate()
    if not data_inicio:
        wd = today.weekday()
        data_inicio = (today - timedelta(days=wd)).isoformat()
    if not data_fim:
        data_fim = (datetime.strptime(data_inicio, '%Y-%m-%d').date() + timedelta(days=6)).isoformat()
    try:
        di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        df = datetime.strptime(data_fim, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse('Parâmetros data_inicio e data_fim inválidos (YYYY-MM-DD).', status=400)
    colaboradores = User.objects.filter(profile__isnull=False).distinct().order_by('profile__nome')
    cid = request.GET.get('colaborador')
    if cid:
        colaboradores = colaboradores.filter(pk=cid)
    linhas = construir_espelho(di, df, colaboradores)
    # Ordenação cronológica para exportação (dia 1 -> dia 31)
    linhas.sort(key=lambda x: (x['data'], (x['nome'] or '').lower()))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Espelho de Ponto'
    headers = ['Data', 'Colaborador', 'Entrada', 'Almoço', 'Volta', 'Saída', 'Total']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
        ws.cell(row=1, column=c).font = Font(bold=True)
    for row_idx, ln in enumerate(linhas, 2):
        marcs = ln.get('marcacoes', [])
        def _hora(i, _m=marcs):
            return _m[i]['hora'] if i < len(_m) else '—'
        ws.cell(row=row_idx, column=1, value=ln['data_str'])
        ws.cell(row=row_idx, column=2, value=ln['nome'])
        ws.cell(row=row_idx, column=3, value=_hora(0))
        ws.cell(row=row_idx, column=4, value=_hora(1))
        ws.cell(row=row_idx, column=5, value=_hora(2))
        ws.cell(row=row_idx, column=6, value=_hora(3))
        ws.cell(row=row_idx, column=7, value=ln['total'])
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="espelho_ponto_{data_inicio}_{data_fim}.xlsx"'
    return response


@backoffice_required
@require_GET
def export_espelho_pdf_view(request):
    """Exporta espelho de ponto para PDF (ReportLab)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
    except ImportError:
        return HttpResponse('reportlab não instalado. pip install reportlab', status=500)
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    today = timezone.localdate()
    if not data_inicio:
        wd = today.weekday()
        data_inicio = (today - timedelta(days=wd)).isoformat()
    if not data_fim:
        data_fim = (datetime.strptime(data_inicio, '%Y-%m-%d').date() + timedelta(days=6)).isoformat()
    try:
        di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        df = datetime.strptime(data_fim, '%Y-%m-%d').date()
    except ValueError:
        return HttpResponse('Parâmetros data_inicio e data_fim inválidos.', status=400)
    User = __import__('django.contrib.auth', fromlist=['get_user_model']).get_user_model()
    colaboradores = User.objects.filter(profile__isnull=False).distinct().order_by('profile__nome')
    cid = request.GET.get('colaborador')
    
    if not cid:
        return HttpResponse('É necessário selecionar um colaborador para exportar o PDF individual.', status=400)
        
    colaboradores = colaboradores.filter(pk=cid)
    linhas = construir_espelho(di, df, colaboradores)
    # Ordenação cronológica para exportação (dia 1 -> dia 31)
    linhas.sort(key=lambda x: (x['data'], (x['nome'] or '').lower()))
    
    from io import BytesIO
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1*cm, leftMargin=1*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    
    # Estilos customizados
    header_info_style = styles['Normal']
    header_info_style.fontSize = 11
    
    # Elementos do PDF
    elements = []
    
    # Info do Funcionário e Mês
    nome_colab = "Colaborador não encontrado"
    mes_ref = di.strftime('%m/%Y')
    if colaboradores.exists():
        nome_colab = colaboradores.first().profile.nome or colaboradores.first().email
        
    elements.append(Paragraph(f"<b>Funcionário:</b> {nome_colab} &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; <b>Mês:</b> {mes_ref}", header_info_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Definição da Tabela (2 níveis de cabeçalho)
    table_data = [
        ['', '1º PERÍODO', '', '2º PERÍODO', '', '', ''],
        ['Data', 'Entrada / Rubrica', 'Saída / Rubrica', 'Entrada / Rubrica', 'Saída / Rubrica', 'Total Horas', 'Observações']
    ]
    
    for ln in linhas:
        marcs = ln.get('marcacoes', [])
        def _get_hora(idx, _marcs=marcs):
            return _marcs[idx]['hora'] if idx < len(_marcs) else ''
        table_data.append([
            ln['data_str'],
            _get_hora(0),
            _get_hora(1),
            _get_hora(2),
            _get_hora(3),
            ln['total'],
            ''  # Observações vazia para preenchimento
        ])
    
    # Larguras
    col_widths = [2.2*cm, 2.7*cm, 2.7*cm, 2.7*cm, 2.7*cm, 2.2*cm, 3.8*cm]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=2)
    
    t.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('SPAN', (1, 0), (2, 0)),
        ('SPAN', (3, 0), (4, 0)),
        ('BACKGROUND', (0, 0), (-1, 1), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 1), 6),
        ('TOPPADDING', (0, 0), (-1, 1), 6),
    ]))
    
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    # Content-Disposition: inline permite pré-visualização no navegador
    response['Content-Disposition'] = f'inline; filename="espelho_ponto_{data_inicio}_{data_fim}.pdf"'
    return response


# ---------- Viaturas ----------
@backoffice_required
def viatura_list_view(request):
    """Lista viaturas com modal para novo."""
    from .forms import ViaturaForm
    viaturas = Viatura.objects.all().order_by('matricula')
    form = ViaturaForm()
    return render(request, 'backoffice/viatura_list.html', {
        'viaturas': viaturas,
        'form': form
    })


@backoffice_required
def viatura_create_view(request):
    """Formulário para cadastrar nova viatura."""
    from .forms import ViaturaForm
    if request.method == 'POST':
        form = ViaturaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('backoffice_viatura_list')
    else:
        form = ViaturaForm()
    return render(request, 'backoffice/viatura_form.html', {'form': form, 'titulo': 'Nova viatura'})


@backoffice_required
def viatura_edit_view(request, pk):
    """Formulário para editar viatura."""
    from .forms import ViaturaForm
    viatura = get_object_or_404(Viatura, pk=pk)
    if request.method == 'POST':
        form = ViaturaForm(request.POST, instance=viatura)
        if form.is_valid():
            form.save()
            return redirect('backoffice_viatura_list')
    else:
        form = ViaturaForm(instance=viatura)
        
    atribuicoes = viatura.historico_atribuicoes.select_related('colaborador__profile').all()
    registos_km = viatura.registos.select_related('utilizador__profile').all().order_by('-timestamp')
    abastecimentos = viatura.abastecimentos.select_related('utilizador__profile').all().order_by('-timestamp')
    
    return render(request, 'backoffice/viatura_form.html', {
        'form': form, 
        'titulo': 'Editar viatura', 
        'viatura': viatura,
        'atribuicoes': atribuicoes,
        'registos_km': registos_km,
        'abastecimentos': abastecimentos
    })


# ---------- Cartões de Combustível ----------
@backoffice_required
def cartao_list_view(request):
    """Lista cartões com modal para novo."""
    from .forms import CartaoForm
    from .models import Cartao
    cartoes = Cartao.objects.all().order_by('nome')
    form = CartaoForm()
    return render(request, 'backoffice/cartao_list.html', {
        'cartoes': cartoes,
        'form': form
    })


@backoffice_required
def cartao_create_view(request):
    """Formulário para cadastrar novo cartão."""
    from .forms import CartaoForm
    from .models import Cartao
    if request.method == 'POST':
        form = CartaoForm(request.POST)
        if form.is_valid():
            cartao_inst = form.save()
            # Se foi atribuído um colaborador ao cartão, atualiza o perfil dele
            colab = form.cleaned_data.get('colaborador_atual')
            if colab and hasattr(colab, 'profile'):
                colab.profile.cartao = cartao_inst
                colab.profile.save()
            return redirect('backoffice_cartao_list')
    else:
        form = CartaoForm()
    return render(request, 'backoffice/cartao_form.html', {'form': form, 'titulo': 'Novo Cartão de Combustível'})


@backoffice_required
def cartao_edit_view(request, pk):
    """Formulário para editar cartão."""
    from .forms import CartaoForm
    from .models import Cartao
    cartao_inst = get_object_or_404(Cartao, pk=pk)
    old_colab = cartao_inst.colaborador_atual
    
    if request.method == 'POST':
        form = CartaoForm(request.POST, instance=cartao_inst)
        if form.is_valid():
            cartao_inst = form.save()
            
            # Limpa o cartão do colaborador antigo se mudou
            colab = form.cleaned_data.get('colaborador_atual')
            if old_colab and old_colab != colab and hasattr(old_colab, 'profile') and old_colab.profile.cartao == cartao_inst:
                old_colab.profile.cartao = None
                old_colab.profile.save()
                
            # Associa ao colaborador novo
            if colab and hasattr(colab, 'profile'):
                colab.profile.cartao = cartao_inst
                colab.profile.save()
                
            return redirect('backoffice_cartao_list')
    else:
        form = CartaoForm(instance=cartao_inst)
        
    atribuicoes = cartao_inst.historico_atribuicoes.select_related('colaborador__profile').all()
    abastecimentos = cartao_inst.abastecimentos.select_related('utilizador__profile', 'viatura').all().order_by('-timestamp')
    
    return render(request, 'backoffice/cartao_form.html', {
        'form': form, 
        'titulo': 'Editar Cartão', 
        'cartao': cartao_inst,
        'atribuicoes': atribuicoes,
        'abastecimentos': abastecimentos
    })


from .models import Abastecimento

@backoffice_required
def backoffice_abastecimento_list_view(request):
    """Listagem de abastecimentos no Backoffice com filtros de status e totais do mês."""
    status_filter = request.GET.get('status')
    
    abastecimentos = Abastecimento.objects.all().order_by('-timestamp')
    if status_filter in ['pendente', 'aprovado', 'rejeitado']:
        abastecimentos = abastecimentos.filter(status=status_filter)
        
    # Calcular estatísticas do mês atual
    from django.utils import timezone
    from django.db.models import Sum
    agora = timezone.localtime(timezone.now())
    primeiro_dia = agora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    total_gasto_mes = Abastecimento.objects.filter(
        status='aprovado',
        timestamp__gte=primeiro_dia
    ).aggregate(Sum('valor'))['valor__sum'] or 0.0
    
    total_pendente_mes = Abastecimento.objects.filter(
        status='pendente'
    ).count()

    total_litros_mes = Abastecimento.objects.filter(
        status='aprovado',
        timestamp__gte=primeiro_dia
    ).aggregate(Sum('litros'))['litros__sum'] or 0.0

    return render(request, 'backoffice/abastecimento_list.html', {
        'abastecimentos': abastecimentos,
        'status_filter': status_filter,
        'total_gasto_mes': total_gasto_mes,
        'total_pendente_mes': total_pendente_mes,
        'total_litros_mes': total_litros_mes
    })


@backoffice_required
def backoffice_abastecimento_aprovar_view(request, pk):
    """Aprovar um abastecimento."""
    if request.method == 'POST':
        abast = get_object_or_404(Abastecimento, pk=pk)
        abast.status = 'aprovado'
        abast.save()
    return redirect('backoffice_abastecimento_list')


@backoffice_required
def backoffice_abastecimento_rejeitar_view(request, pk):
    """Rejeitar um abastecimento indicando justificativa."""
    if request.method == 'POST':
        abast = get_object_or_404(Abastecimento, pk=pk)
        justificativa = request.POST.get('justificativa_admin')
        if not justificativa:
            justificativa = "Rejeitado pela administração."
        abast.status = 'rejeitado'
        abast.justificativa_admin = justificativa
        abast.save()
    return redirect('backoffice_abastecimento_list')


def _obter_dados_historico(request):
    from .models import Viatura, Cartao, ViaturaHistoricoAtribuicao, CartaoHistoricoAtribuicao, Abastecimento, RegistroKM
    from django.contrib.auth import get_user_model
    from datetime import datetime
    from django.utils import timezone
    from django.utils.timezone import make_aware
    from django.db.models import Q
    
    User = get_user_model()
    
    data_inicio_str = request.GET.get('data_inicio')
    data_fim_str = request.GET.get('data_fim')
    viatura_id = request.GET.get('viatura')
    cartao_id = request.GET.get('cartao')
    colaborador_id = request.GET.get('colaborador')
    
    today = timezone.localdate()
    if not data_inicio_str:
        data_inicio = today.replace(day=1)
    else:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        
    if not data_fim_str:
        data_fim = today
    else:
        data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        
    start_dt = make_aware(datetime.combine(data_inicio, datetime.min.time()))
    end_dt = make_aware(datetime.combine(data_fim, datetime.max.time()))
    
    atrib_viaturas = ViaturaHistoricoAtribuicao.objects.select_related('viatura', 'colaborador__profile').filter(
        data_atribuicao__lte=end_dt
    ).filter(Q(data_remocao__isnull=True) | Q(data_remocao__gte=start_dt))
    
    atrib_cartoes = CartaoHistoricoAtribuicao.objects.select_related('cartao', 'colaborador__profile').filter(
        data_atribuicao__lte=end_dt
    ).filter(Q(data_remocao__isnull=True) | Q(data_remocao__gte=start_dt))
    
    abastecimentos = Abastecimento.objects.select_related('viatura', 'utilizador__profile', 'cartao_ref').filter(
        timestamp__range=(start_dt, end_dt)
    )
    
    registos_km = RegistroKM.objects.select_related('viatura', 'utilizador__profile').filter(
        timestamp__range=(start_dt, end_dt)
    )
    
    if viatura_id:
        atrib_viaturas = atrib_viaturas.filter(viatura_id=viatura_id)
        abastecimentos = abastecimentos.filter(viatura_id=viatura_id)
        registos_km = registos_km.filter(viatura_id=viatura_id)
        if not cartao_id:
            atrib_cartoes = atrib_cartoes.none()
            
    if cartao_id:
        atrib_cartoes = atrib_cartoes.filter(cartao_id=cartao_id)
        abastecimentos = abastecimentos.filter(cartao_ref_id=cartao_id)
        if not viatura_id:
            atrib_viaturas = atrib_viaturas.none()
            registos_km = registos_km.none()
            
    if colaborador_id:
        atrib_viaturas = atrib_viaturas.filter(colaborador_id=colaborador_id)
        atrib_cartoes = atrib_cartoes.filter(colaborador_id=colaborador_id)
        abastecimentos = abastecimentos.filter(utilizador_id=colaborador_id)
        registos_km = registos_km.filter(utilizador_id=colaborador_id)
        
    consolidated_logs = []
    
    for av in atrib_viaturas:
        colab_name = av.colaborador.profile.nome or av.colaborador.email
        status_str = "Ativo" if not av.data_remocao else f"até {timezone.localtime(av.data_remocao).strftime('%d/%m/%Y %H:%M')}"
        consolidated_logs.append({
            'tipo': 'atribuicao_viatura',
            'tipo_label': 'Atribuição Viatura',
            'data': av.data_atribuicao,
            'colaborador': colab_name,
            'recurso': f"Viatura {av.viatura.matricula}",
            'descricao': f"Viatura {av.viatura.matricula} atribuída a {colab_name} ({status_str})",
            'detalhes': f"Início: {timezone.localtime(av.data_atribuicao).strftime('%d/%m/%Y %H:%M')} | Fim: {status_str}"
        })
        
    for ac in atrib_cartoes:
        colab_name = ac.colaborador.profile.nome or ac.colaborador.email
        status_str = "Ativo" if not ac.data_remocao else f"até {timezone.localtime(ac.data_remocao).strftime('%d/%m/%Y %H:%M')}"
        consolidated_logs.append({
            'tipo': 'atribuicao_cartao',
            'tipo_label': 'Atribuição Cartão',
            'data': ac.data_atribuicao,
            'colaborador': colab_name,
            'recurso': f"Cartão {ac.cartao.nome} ({ac.cartao.numero[-4:] if len(ac.cartao.numero) >= 4 else ac.cartao.numero})",
            'descricao': f"Cartão {ac.cartao.nome} atribuído a {colab_name} ({status_str})",
            'detalhes': f"Início: {timezone.localtime(ac.data_atribuicao).strftime('%d/%m/%Y %H:%M')} | Fim: {status_str}"
        })
        
    for ab in abastecimentos:
        colab_name = ab.utilizador.profile.nome or ab.utilizador.email
        viatura_str = f"Viatura {ab.viatura.matricula}" if ab.viatura else "Sem Viatura"
        pagamento_str = "Cartão" if ab.metodo_pagamento == 'cartao' else "Dinheiro (€)"
        consolidated_logs.append({
            'tipo': 'abastecimento',
            'tipo_label': 'Abastecimento',
            'data': ab.timestamp,
            'colaborador': colab_name,
            'recurso': viatura_str,
            'descricao': f"Abastecimento na {viatura_str} via {pagamento_str}: {ab.litros or 0}L (€{ab.valor})",
            'detalhes': f"Valor: €{ab.valor} | Volume: {ab.litros or 0}L | Cartão: {ab.cartao or 'Não aplicado'} | Status: {ab.get_status_display()}"
        })
        
    for km in registos_km:
        colab_name = km.utilizador.profile.nome or km.utilizador.email
        consolidated_logs.append({
            'tipo': 'registro_km',
            'tipo_label': 'Registo de KM',
            'data': km.timestamp,
            'colaborador': colab_name,
            'recurso': f"Viatura {km.viatura.matricula}",
            'descricao': f"Registo de KM na Viatura {km.viatura.matricula}: {km.km} KM",
            'detalhes': f"Leitura de KM: {km.km} KM"
        })
        
    consolidated_logs.sort(key=lambda x: x['data'], reverse=True)
    return consolidated_logs, data_inicio, data_fim


@backoffice_required
def backoffice_historico_view(request):
    from .models import Viatura, Cartao
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    viatura_id = request.GET.get('viatura')
    cartao_id = request.GET.get('cartao')
    colaborador_id = request.GET.get('colaborador')
    
    consolidated_logs, data_inicio, data_fim = _obter_dados_historico(request)
    
    tab_atribuicoes = [log for log in consolidated_logs if log['tipo'] in ['atribuicao_viatura', 'atribuicao_cartao']]
    tab_abastecimentos = [log for log in consolidated_logs if log['tipo'] == 'abastecimento']
    tab_km = [log for log in consolidated_logs if log['tipo'] == 'registro_km']
    
    viaturas_list = Viatura.objects.filter(ativo=True).order_by('matricula')
    cartoes_list = Cartao.objects.filter(ativo=True).order_by('nome')
    colaboradores_list = User.objects.filter(profile__isnull=False).distinct().order_by('profile__nome')
    
    return render(request, 'backoffice/historico_dashboard.html', {
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'viatura_selecionada': viatura_id,
        'cartao_selecionado': cartao_id,
        'colaborador_selecionado': colaborador_id,
        
        'viaturas': viaturas_list,
        'cartoes': cartoes_list,
        'colaboradores': colaboradores_list,
        
        'consolidated_logs': consolidated_logs,
        'tab_atribuicoes': tab_atribuicoes,
        'tab_abastecimentos': tab_abastecimentos,
        'tab_km': tab_km
    })


@backoffice_required
def export_historico_excel_view(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return HttpResponse('openpyxl não está instalado.', status=500)
        
    logs, di, df = _obter_dados_historico(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    
    # Styling definitions
    font_title = Font(name="Arial", size=14, bold=True, color="1E293B")
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="Arial", size=10, color="000000")
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Relatório de Auditoria e Histórico ({di.strftime('%d/%m/%Y')} - {df.strftime('%d/%m/%Y')})"
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    ws.row_dimensions[1].height = 30
    
    # Blank row
    ws.append([])
    
    # Headers
    headers = ["Data/Hora", "Tipo de Registo", "Colaborador", "Recurso", "Descrição", "Detalhes Adicionais"]
    ws.append(headers)
    ws.row_dimensions[3].height = 24
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    # Data Rows
    from django.utils import timezone
    for log in logs:
        local_dt = timezone.localtime(log['data'])
        dt_str = local_dt.strftime('%d/%m/%Y %H:%M:%S')
        
        row_data = [
            dt_str,
            log['tipo_label'],
            log['colaborador'],
            log['recurso'],
            log['descricao'],
            log['detalhes']
        ]
        ws.append(row_data)
        
    # Styling and auto-fit columns
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
        ws.row_dimensions[row[0].row].height = 20
        for cell in row:
            cell.font = font_body
            cell.border = thin_border
            if cell.column in [1, 2]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
                
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="historico_auditoria_{di}_{df}.xlsx"'
    wb.save(response)
    return response


@backoffice_required
def export_historico_pdf_view(request):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from io import BytesIO
    except ImportError:
        return HttpResponse('reportlab não está instalado.', status=500)
        
    logs, di, df = _obter_dados_historico(request)
    
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1.2*cm,
        bottomMargin=1.2*cm
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=16,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=15
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1
    )
    
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#334155')
    )
    
    cell_center_style = ParagraphStyle(
        'CellCenterStyle',
        parent=cell_style,
        alignment=1
    )
    
    elements = []
    
    elements.append(Paragraph(f"Relatório de Auditoria e Histórico ({di.strftime('%d/%m/%Y')} - {df.strftime('%d/%m/%Y')})", title_style))
    elements.append(Spacer(1, 10))
    
    table_data = [[
        Paragraph("Data/Hora", header_style),
        Paragraph("Tipo de Registo", header_style),
        Paragraph("Colaborador", header_style),
        Paragraph("Recurso", header_style),
        Paragraph("Descrição", header_style),
        Paragraph("Detalhes Adicionais", header_style)
    ]]
    
    from django.utils import timezone
    for log in logs:
        local_dt = timezone.localtime(log['data'])
        dt_str = local_dt.strftime('%d/%m/%Y %H:%M')
        
        table_data.append([
            Paragraph(dt_str, cell_center_style),
            Paragraph(log['tipo_label'], cell_center_style),
            Paragraph(log['colaborador'], cell_style),
            Paragraph(log['recurso'], cell_style),
            Paragraph(log['descricao'], cell_style),
            Paragraph(log['detalhes'], cell_style)
        ])
        
    col_widths = [95, 95, 110, 110, 225, 150]
    
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    for i in range(1, len(table_data)):
        bg_color = colors.HexColor('#F8FAFC') if i % 2 == 0 else colors.white
        t.setStyle(TableStyle([('BACKGROUND', (0, i), (-1, i), bg_color)]))
        
    elements.append(t)
    doc.build(elements)
    
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="historico_auditoria_{di}_{df}.pdf"'
    return response
